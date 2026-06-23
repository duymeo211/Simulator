#!/usr/bin/env python3
"""Generate per-channel DBC files from the 740D Multi-Media CAN-FD spec.

The 'Transmit(FD)' sheet is the master bit-assignment table for every message
(CAN ID, DLC, and Motorola/big-endian signal layout). Columns 75-106 mark which
of the three buses each message belongs to. We split into one DBC per bus.
"""
import re
from collections import Counter
from openpyxl import load_workbook

SRC = "/mnt/user-data/uploads/24CY_Multi-Media_system_CAN_Com_spec_v7_30_BEVstep3_BA.xlsx"
OUTDIR = "/mnt/user-data/outputs"

# column indices (0-based) in the master Transmit(FD) sheet
C_LABEL, C_VAR, C_NAME_E, C_NAME_J = 0, 1, 2, 3
C_ID, C_FDF, C_DLC = 4, 5, 6
C_CYCLE = 8
C_POS, C_LEN, C_DLABEL, C_DNAME_E = 172, 173, 174, 175
C_SIGN, C_UNIT, C_RES, C_OFF = 184, 185, 186, 187

BUS_RANGES = {           # bus -> inclusive column range that marks membership
    "G2M-1": (75, 86),
    "G2M-2": (87, 98),
    "G5M":   (99, 106),
}
BUS_FILE = {
    "G2M-1": "740D_CANFD_G2M-1.dbc",
    "G2M-2": "740D_CANFD_G2M-2.dbc",
    "G5M":   "740D_CANFD_G5M.dbc",
}
NODE = "MM"  # the Multi-Media head unit this spec belongs to

# Signal rows whose Data-Label cell carries one of these fills are ignored.
# In this workbook's custom palette: grey = 808080 (idx 23), purple = 403151 (idx 28).
# They mark redundant variation-alternate / redacted-placeholder duplicate rows.
DROP_FILL_RGB = {"808080", "403151"}  # last-6 hex (grey, purple)


def fill_rgb6(cell, palette):
    """Return the cell's solid-fill colour as last-6 uppercase hex, or None."""
    f = cell.fill
    if not f or f.patternType != "solid" or f.fgColor is None:
        return None
    fg = f.fgColor
    argb = None
    if fg.type == "indexed":
        try:
            argb = palette[fg.indexed]
        except (IndexError, TypeError):
            return None
    elif fg.type == "rgb":
        argb = fg.rgb
    if not isinstance(argb, str):
        return None
    return argb[-6:].upper()


def num(s, default):
    """Parse a numeric cell that may carry notes ('1 see#-', '1/10', '0.940')."""
    if s is None:
        return default
    t = str(s).strip()
    if t in ("", "-", "*"):
        return default
    tok = t.split()[0]
    try:
        if "/" in tok:
            a, b = tok.split("/")
            return float(a) / float(b)
        return float(tok)
    except (ValueError, ZeroDivisionError):
        return default


def parse_pos(s):
    """'Byte' or 'Byte.Bit' -> the signal's LSB as a DBC bit number.

    Numbering convention (per spec):
      * Bytes counted from 1, left to right (byte 1 is the first byte).
      * Bits counted from 0, right to left within a byte (bit 0 = LSB on the
        right, bit 7 = MSB on the left). Bit omitted => 0.
    This byte.bit is the signal's LSB:  lsb = (byte - 1) * 8 + bit
    The bus is big-endian (Motorola), so the signal extends from this LSB into
    the lower (more significant) bytes; the DBC start bit is the MSB, derived
    from this LSB by lsb_to_msb().  e.g. I_PMD "2.0" len 10 -> LSB bit 8 ->
    MSB bit 1 -> occupies byte 1 bits 0-1 + byte 2 bits 0-7.
    """
    t = str(s).strip()
    if "." in t:
        b, bit = t.split(".", 1)
        byte, off = int(b), int(bit)
    else:
        byte, off = int(t), 0
    return (byte - 1) * 8 + off


def sanitize(name, fallback):
    name = re.sub(r"[^A-Za-z0-9_]", "_", str(name).strip())
    if not name or not re.match(r"[A-Za-z_]", name):
        name = fallback if re.match(r"[A-Za-z_]", str(fallback)) else "S_" + name
    return name


def fmt(x):
    if x == int(x):
        return str(int(x))
    return ("%.12f" % x).rstrip("0").rstrip(".")


def load_messages():
    # Not read_only: we need cell fill colours to filter grey/purple rows.
    wb = load_workbook(SRC, data_only=True)
    palette = wb._colors                      # workbook's custom indexed palette
    ws = wb["Transmit(FD)"]
    messages, cur = [], None
    color_dropped = [0]

    def cv(row, i):
        return row[i].value if i < len(row) else None

    for row in ws.iter_rows(min_row=15):
        if cv(row, C_ID) is not None and cv(row, C_LABEL) is not None:
            try:
                cid = int(str(cv(row, C_ID)).strip(), 16)
            except ValueError:
                cur = None
                continue
            try:
                dlc = int(str(cv(row, C_DLC)).strip())
            except (ValueError, TypeError):
                dlc = 8
            cur = {
                "label": str(cv(row, C_LABEL)).strip(),
                "var": "" if cv(row, C_VAR) is None else str(cv(row, C_VAR)).strip(),
                "name": "" if cv(row, C_NAME_E) is None else str(cv(row, C_NAME_E)).strip(),
                "id": cid, "dlc": dlc,
                "fd": str(cv(row, C_FDF)).strip() in ("1", "1.0"),
                "cycle": int(num(cv(row, C_CYCLE), 0)),
                "buses": [b for b, (lo, hi) in BUS_RANGES.items()
                          if any(cv(row, c) not in (None, "") for c in range(lo, hi + 1))],
                "signals": [],
            }
            messages.append(cur)
        elif cv(row, C_POS) is not None and cur is not None:
            # Skip signal rows flagged grey/purple in the spec.
            if fill_rgb6(row[C_DLABEL], palette) in DROP_FILL_RGB:
                color_dropped[0] += 1
                continue
            try:
                start = parse_pos(cv(row, C_POS))
                length = int(str(cv(row, C_LEN)).strip())
            except (ValueError, TypeError):
                continue
            factor = num(cv(row, C_RES), 1.0) or 1.0
            unit = str(cv(row, C_UNIT)).strip()
            cur["signals"].append({
                "name": "" if cv(row, C_DLABEL) is None else str(cv(row, C_DLABEL)).strip(),
                "desc": "" if cv(row, C_DNAME_E) is None else str(cv(row, C_DNAME_E)).strip(),
                "start": start,                          # spec LSB bit (see parse_pos)
                "len": length,
                "signed": str(cv(row, C_SIGN)).strip().upper() == "S",
                "factor": factor,
                "offset": num(cv(row, C_OFF), 0.0),
                "unit": "" if unit in ("-", "None", "") else unit,
            })
    print("Ignored %d grey/purple signal rows" % color_dropped[0])
    return messages


def motorola_bits(start, length):
    """Bit indices consumed by a big-endian (Motorola) signal, MSB at `start`."""
    bits, pos = [], start
    for _ in range(length):
        bits.append(pos)
        pos = pos + 15 if pos % 8 == 0 else pos - 1
    return bits


def lsb_to_msb(lsb, length):
    """Given a signal's LSB bit number, return its MSB (DBC Motorola start bit).

    Big-endian fill: from the LSB, consume bits upward within the byte
    (..0,1,2..7) then cross into the next lower byte address (more significant).
    The last bit consumed is the MSB.
    """
    byte, pos = divmod(lsb, 8)
    msb = lsb
    for _ in range(length):
        msb = byte * 8 + pos
        if pos == 7:
            pos, byte = 0, byte - 1
        else:
            pos += 1
    return msb


def build_signal_line(s, used_names, dlc):
    # s["start"] is the spec LSB; the bus is big-endian, so the DBC start bit
    # is the MSB. Validate the whole occupied range fits the frame.
    if s["len"] <= 0 or s["start"] < 0:
        return None
    start = lsb_to_msb(s["start"], s["len"])
    bits = motorola_bits(start, s["len"])
    if min(bits) < 0 or max(bits) >= dlc * 8:
        return None  # signal does not physically fit the frame -> drop it
    name = sanitize(s["name"], "Sig")
    base, i = name, 2
    while name in used_names:
        name = "%s_%d" % (base, i); i += 1
    used_names.add(name)
    bo = 0  # Motorola / big-endian for all signals in this spec
    sign = "-" if s["signed"] else "+"
    f, off = s["factor"], s["offset"]
    if s["signed"]:
        rmin, rmax = -(2 ** (s["len"] - 1)), 2 ** (s["len"] - 1) - 1
    else:
        rmin, rmax = 0, 2 ** s["len"] - 1
    p1, p2 = rmin * f + off, rmax * f + off
    lo, hi = (p1, p2) if p1 <= p2 else (p2, p1)
    return ' SG_ %s : %d|%d@%d%s (%s,%s) [%s|%s] "%s" %s' % (
        name, start, s["len"], bo, sign,
        fmt(f), fmt(off), fmt(lo), fmt(hi), s["unit"], NODE)


HEADER = '''VERSION ""

NS_ :
\tNS_DESC_
\tCM_
\tBA_DEF_
\tBA_
\tVAL_
\tCAT_DEF_
\tCAT_
\tFILTER
\tBA_DEF_DEF_
\tEV_DATA_
\tENVVAR_DATA_
\tSGTYPE_
\tSGTYPE_VAL_
\tBA_DEF_SGTYPE_
\tBA_SGTYPE_
\tSIG_TYPE_REF_
\tVAL_TABLE_
\tSIG_GROUP_
\tSIG_VALTYPE_
\tSIGTYPE_VALTYPE_
\tBO_TX_BU_
\tBA_DEF_REL_
\tBA_REL_
\tBA_DEF_DEF_REL_
\tBU_SG_REL_
\tBU_EV_REL_
\tBU_BO_REL_
\tSG_MUL_VAL_

BS_:

BU_: %s
''' % NODE


def write_bus(bus, messages):
    lines = [HEADER]
    seen_ids, used_msg_names, attrs, comments, dropped = set(), set(), [], [], []
    dropped_sigs = [0]
    for m in messages:
        if bus not in m["buses"]:
            continue
        if m["id"] in seen_ids:
            dropped.append((m["id"], m["label"], m["var"]))
            continue
        seen_ids.add(m["id"])
        mname = sanitize(m["label"], "MSG_%X" % m["id"])
        base, i = mname, 2
        while mname in used_msg_names:
            suf = m["var"] or str(i)
            mname = "%s_%s" % (base, sanitize(suf, str(i)))
            if mname in used_msg_names:
                mname = "%s_%d" % (base, i)
            i += 1
        used_msg_names.add(mname)
        lines.append("\nBO_ %d %s: %d %s" % (m["id"], mname, m["dlc"], NODE))
        used_sig = set()
        for s in sorted(m["signals"], key=lambda x: lsb_to_msb(x["start"], x["len"])):
            sl = build_signal_line(s, used_sig, m["dlc"])
            if sl is None:
                dropped_sigs[0] += 1
                continue
            lines.append(sl)
        if m["name"]:
            comments.append('CM_ BO_ %d "%s";' % (m["id"], m["name"].replace('"', "'")))
        attrs.append('BA_ "VFrameFormat" BO_ %d %d;' % (m["id"], 4 if m["fd"] else 0))
        if m["cycle"] > 0:
            attrs.append('BA_ "GenMsgCycleTime" BO_ %d %d;' % (m["id"], m["cycle"]))

    lines.append("\n")
    lines.append('BA_DEF_ BO_  "VFrameFormat" ENUM  "StandardCAN","ExtendedCAN",'
                 '"reserved","J1939PG","StandardCAN_FD","ExtendedCAN_FD";')
    lines.append('BA_DEF_ BO_  "GenMsgCycleTime" INT 0 65535;')
    lines.append('BA_DEF_DEF_  "VFrameFormat" "StandardCAN";')
    lines.append('BA_DEF_DEF_  "GenMsgCycleTime" 0;')
    lines.extend(attrs)
    lines.append(
        'CM_ "740D Multi-Media CAN-FD - bus CANFD_%s_BUS. '
        'Source: gncanvehcs spec v0003-a-20250402. '
        'Signals: byte-1-indexed, bit0=LSB (right), Data Pos = LSB, big-endian (Motorola). '
        'Checksum/counter (KZK*), full-frame (FV*), redacted template (XYZ*/D#####) and '
        'oversized meta signals omitted; duplicate-CAN-ID variations merged to first.";'
        % bus.replace("-", "_"))
    lines.extend(comments)
    text = "\n".join(lines) + "\n"
    path = "%s/%s" % (OUTDIR, BUS_FILE[bus])
    with open(path, "w", encoding="utf-8") as f:
        f.write(text)
    return path, len(seen_ids), dropped, dropped_sigs[0]


def main():
    import os
    os.makedirs(OUTDIR, exist_ok=True)
    messages = load_messages()
    print("Parsed %d messages, %d signals total"
          % (len(messages), sum(len(m["signals"]) for m in messages)))
    for bus in BUS_RANGES:
        path, n, dropped, nsig_drop = write_bus(bus, messages)
        print("\n%-6s -> %s : %d messages, %d placeholder/oversized signals dropped"
              % (bus, os.path.basename(path), n, nsig_drop))
        if dropped:
            print("   %d duplicate-ID variation(s) merged-out (kept first):" % len(dropped))
            for cid, lab, var in dropped[:6]:
                print("     0x%03X %s/%s" % (cid, lab, var))
            if len(dropped) > 6:
                print("     ...")


if __name__ == "__main__":
    main()
